import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def save_to_excel(name, age, gender, symptoms, diagnosis, urgency):
    filename = "diagnosis_log.xlsx"
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "Gender", "Symptoms", "Diagnosis", "Urgency"])

    ws.append([name, age, gender, symptoms, diagnosis, urgency])
    row = ws.max_row

    # Color-code urgency
    color_map = {
        "Critical": "FF0000",  # Red
        "Urgent": "FFA500",   # Orange
        "Normal": "008000"    # Green
    }
    color = color_map.get(urgency, "FFFFFF")

    for cell in ws[row]:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    wb.save(filename)
